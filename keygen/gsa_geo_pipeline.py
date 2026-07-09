#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified GSA/A-Parser query pipeline v1.

Input workbook: gsa_geo_pipeline_keywords_v1.xlsx
Required sheets:
  - Keywords_Pipeline
  - Footprint_Families

Output query format per line:
  {Seed} {Operator} {Footprint}

Key design:
  - one all-geo pool, no hard P1/P2/P3 cluster split
  - only soft KPI weights: 1.05-1.25
  - strict geo -> ccTLD operator binding
  - small query batches, default 6 MB instead of 40 MB
  - rotate a subset of footprints per seed per batch, default 24
  - do not stop whole pool when one geo/source is exhausted
  - pages 20/25/30 are A-Parser project settings; they are recorded in manifest

Dependencies:
  pip install openpyxl

Example:
  python gsa_geo_pipeline.py \
    --input-xlsx gsa_geo_pipeline_keywords_v1.xlsx \
    --out-dir parser_batches \
    --target-mb 6 \
    --pages 25 \
    --footprints-per-seed 24 \
    --batches 10
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import random
import re
import sqlite3
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

OUTPUT_ENCODING = "utf-8"
DEFAULT_TARGET_MB = 6
DEFAULT_PAGES = 25
DEFAULT_FOOTPRINTS_PER_SEED = 80
SHUFFLE_BUFFER_LINES = 50000


@dataclass
class SeedRow:
    pipeline_id: int
    geo_unit: str
    seed: str
    operators: List[str]
    soft_weight: float
    kpi: float
    original_priority: str
    language: str


@dataclass
class FootprintRow:
    footprint_id: int
    family: str
    footprint: str
    enabled: bool


def norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def split_pipe(value: object) -> List[str]:
    if value is None:
        return []
    items = [x.strip() for x in str(value).split("|")]
    return [x for x in items if x]


def load_sheet_rows(path: Path, sheet_name: str) -> List[Dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [norm_header(x) for x in next(rows)]
    except StopIteration:
        return []
    result = []
    for raw in rows:
        if not raw or not any(x is not None and str(x).strip() for x in raw):
            continue
        row = {header[i]: raw[i] if i < len(raw) else None for i in range(len(header))}
        result.append(row)
    wb.close()
    return result


def load_seeds(path: Path) -> List[SeedRow]:
    rows = load_sheet_rows(path, "Keywords_Pipeline")
    seeds: List[SeedRow] = []
    for row in rows:
        seed = str(row.get("seed") or "").strip()
        if not seed:
            continue
        operators = split_pipe(row.get("operators"))
        if not operators:
            continue
        try:
            weight = float(row.get("soft_weight") or 1.0)
        except Exception:
            weight = 1.0
        try:
            kpi = float(row.get("kpi_weekly_target") or 0)
        except Exception:
            kpi = 0.0
        try:
            pid = int(row.get("pipeline_id") or 0)
        except Exception:
            pid = 0
        seeds.append(SeedRow(
            pipeline_id=pid,
            geo_unit=str(row.get("geo_unit") or "").strip(),
            seed=seed,
            operators=operators,
            soft_weight=max(0.1, min(weight, 2.0)),
            kpi=kpi,
            original_priority=str(row.get("original_priority") or "").strip(),
            language=str(row.get("geo_language") or "").strip(),
        ))
    return seeds


def load_footprints(path: Path) -> List[FootprintRow]:
    rows = load_sheet_rows(path, "Footprint_Families")
    fps: List[FootprintRow] = []
    for row in rows:
        enabled = str(row.get("enabled") or "1").strip().lower() not in {"0", "false", "no", "off"}
        fp = str(row.get("footprint") or "").strip()
        if not enabled or not fp:
            continue
        try:
            fid = int(row.get("footprint_id") or 0)
        except Exception:
            fid = 0
        fps.append(FootprintRow(
            footprint_id=fid,
            family=str(row.get("family") or "other_submit").strip(),
            footprint=fp,
            enabled=True,
        ))
    return fps


class DedupeStore:
    def __init__(self, db_path: Optional[Path]):
        self.db_path = db_path
        self.conn = None
        self.duplicates = 0
        if db_path:
            self.conn = sqlite3.connect(str(db_path))
            self.conn.execute("PRAGMA journal_mode=OFF")
            self.conn.execute("PRAGMA synchronous=OFF")
            self.conn.execute("CREATE TABLE IF NOT EXISTS seen (h TEXT PRIMARY KEY)")
            self.conn.commit()

    def add(self, line: str) -> bool:
        if self.conn is None:
            return True
        h = hashlib.sha1(line.encode(OUTPUT_ENCODING, errors="ignore")).hexdigest()
        try:
            self.conn.execute("INSERT INTO seen(h) VALUES (?)", (h,))
            return True
        except sqlite3.IntegrityError:
            self.duplicates += 1
            return False

    def commit(self) -> None:
        if self.conn:
            self.conn.commit()

    def close(self) -> None:
        if self.conn:
            self.conn.commit()
            self.conn.close()
            self.conn = None


def weighted_shuffle(items: Sequence[SeedRow], rnd: random.Random) -> List[SeedRow]:
    # Efraimidis-Spirakis weighted random permutation.
    keyed = []
    for item in items:
        u = max(rnd.random(), 1e-12)
        key = u ** (1.0 / max(item.soft_weight, 0.01))
        keyed.append((key, item))
    keyed.sort(key=lambda x: x[0], reverse=True)
    return [item for _, item in keyed]


def footprints_by_family(footprints: Sequence[FootprintRow]) -> Dict[str, List[FootprintRow]]:
    fams: Dict[str, List[FootprintRow]] = defaultdict(list)
    for fp in footprints:
        fams[fp.family].append(fp)
    return dict(fams)


def select_footprints(
    footprint_families: Dict[str, List[FootprintRow]],
    batch_index: int,
    seed: SeedRow,
    count: int,
    rnd: random.Random,
) -> List[FootprintRow]:
    families = sorted(footprint_families)
    if not families:
        return []
    # Rotate starting family by batch and seed id; this prevents every seed from using the same family mix.
    start = (batch_index + seed.pipeline_id) % len(families)
    ordered_families = families[start:] + families[:start]
    selected: List[FootprintRow] = []
    per_family = max(1, count // min(len(ordered_families), count))
    for fam in ordered_families:
        pool = footprint_families[fam][:]
        rnd.shuffle(pool)
        selected.extend(pool[:per_family])
        if len(selected) >= count:
            break
    if len(selected) < count:
        all_fps = [fp for fam in ordered_families for fp in footprint_families[fam]]
        rnd.shuffle(all_fps)
        seen_ids = {fp.footprint_id for fp in selected}
        for fp in all_fps:
            if fp.footprint_id in seen_ids:
                continue
            selected.append(fp)
            if len(selected) >= count:
                break
    return selected[:count]


def flush_buffer(buffer: List[str], out_fh, rnd: random.Random) -> int:
    if not buffer:
        return 0
    rnd.shuffle(buffer)
    total = 0
    for line in buffer:
        data = (line + "\n").encode(OUTPUT_ENCODING)
        out_fh.write(data.decode(OUTPUT_ENCODING))
        total += len(data)
    buffer.clear()
    return total


def generate_batches(
    seeds: List[SeedRow],
    footprints: List[FootprintRow],
    out_dir: Path,
    target_mb: float,
    pages: int,
    footprints_per_seed: int,
    batches: int,
    dedupe: DedupeStore,
    random_seed: int,
) -> Dict[str, object]:
    rnd = random.Random(random_seed)
    out_dir.mkdir(parents=True, exist_ok=True)
    target_bytes = int(target_mb * 1024 * 1024)
    fp_families = footprints_by_family(footprints)
    manifest = {
        "format_version": "GSA-GEO-PIPELINE-v1",
        "query_format": "{Seed} {Operator} {Footprint}",
        "pages_per_query_aparser_setting": pages,
        "target_mb": target_mb,
        "footprints_per_seed": footprints_per_seed,
        "input_seed_count": len(seeds),
        "input_footprint_count": len(footprints),
        "family_counts": {fam: len(pool) for fam, pool in fp_families.items()},
        "output_files": [],
    }

    for batch_idx in range(1, batches + 1):
        out_path = out_dir / f"B{batch_idx:04d}_ALL_softgeo_p{pages}_{int(target_mb)}mb.txt"
        geo_counts: Dict[str, int] = defaultdict(int)
        family_counts: Dict[str, int] = defaultdict(int)
        lines_written = 0
        lines_deduped = 0
        bytes_flushed = 0
        bytes_buffered = 0
        buffer: List[str] = []
        seed_order = weighted_shuffle(seeds, rnd)
        seed_cursor = 0

        with open(out_path, "w", encoding=OUTPUT_ENCODING, newline="\n") as out_fh:
            while bytes_flushed + bytes_buffered < target_bytes:
                if seed_cursor >= len(seed_order):
                    seed_order = weighted_shuffle(seeds, rnd)
                    seed_cursor = 0
                seed = seed_order[seed_cursor]
                seed_cursor += 1
                operator = rnd.choice(seed.operators)
                selected_fps = select_footprints(fp_families, batch_idx, seed, footprints_per_seed, rnd)
                for fp in selected_fps:
                    line = f"{seed.seed} {operator} {fp.footprint}"
                    if not dedupe.add(line):
                        lines_deduped += 1
                        continue
                    buffer.append(line)
                    lines_written += 1
                    geo_counts[seed.geo_unit] += 1
                    family_counts[fp.family] += 1
                    bytes_buffered += len((line + "\n").encode(OUTPUT_ENCODING))
                if len(buffer) >= SHUFFLE_BUFFER_LINES:
                    bytes_flushed += flush_buffer(buffer, out_fh, rnd)
                    bytes_buffered = 0
                    dedupe.commit()
            bytes_flushed += flush_buffer(buffer, out_fh, rnd)
            dedupe.commit()

        manifest["output_files"].append({
            "file": str(out_path),
            "bytes": bytes_flushed,
            "mb": round(bytes_flushed / 1024 / 1024, 3),
            "lines_written": lines_written,
            "lines_deduped": lines_deduped,
            "geo_query_counts": dict(sorted(geo_counts.items())),
            "family_query_counts": dict(sorted(family_counts.items())),
        })

    manifest["global_query_line_duplicates"] = dedupe.duplicates
    return manifest


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build unified soft-geo GSA/A-Parser query batches")
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--out-dir", default=Path("parser_batches"), type=Path)
    parser.add_argument("--target-mb", default=DEFAULT_TARGET_MB, type=float)
    parser.add_argument("--pages", default=DEFAULT_PAGES, type=int, help="A-Parser pages per query setting to record in manifest")
    parser.add_argument("--footprints-per-seed", default=DEFAULT_FOOTPRINTS_PER_SEED, type=int)
    parser.add_argument("--batches", default=1, type=int)
    parser.add_argument("--random-seed", default=20260601, type=int)
    parser.add_argument("--no-dedupe", action="store_true")
    args = parser.parse_args(argv)

    if args.pages < 1:
        raise ValueError("--pages must be >= 1")
    if args.footprints_per_seed < 1:
        raise ValueError("--footprints-per-seed must be >= 1")

    seeds = load_seeds(args.input_xlsx)
    footprints = load_footprints(args.input_xlsx)
    if not seeds:
        raise RuntimeError("No usable seeds found in Keywords_Pipeline")
    if not footprints:
        raise RuntimeError("No usable footprints found in Footprint_Families")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    dedupe_path = None if args.no_dedupe else args.out_dir / "query_dedupe.sqlite"
    dedupe = DedupeStore(dedupe_path)
    try:
        manifest = generate_batches(
            seeds=seeds,
            footprints=footprints,
            out_dir=args.out_dir,
            target_mb=args.target_mb,
            pages=args.pages,
            footprints_per_seed=args.footprints_per_seed,
            batches=args.batches,
            dedupe=dedupe,
            random_seed=args.random_seed,
        )
    finally:
        dedupe.close()

    manifest_path = args.out_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Done. Files: {len(manifest['output_files'])}. Manifest: {manifest_path}")
    for item in manifest["output_files"]:
        print(f"{item['file']}: {item['mb']} MB, {item['lines_written']} lines, deduped {item['lines_deduped']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
